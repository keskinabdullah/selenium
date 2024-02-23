from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.keys import Keys
import pytest
import openpyxl
from constants import globalConstants as c

#Kullanıcı adı ve şifre alanları boş geçildiğinde uyarı mesajı olarak "Epic sadface: Username is required" gösterilmelidir  

class Test_Sauce_Demo_Login_Test:  
        #prefix => test_ 
    def setup_method(self): #her test başlangıcında çalışacak fonk
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window() 
    
    def teardown_method(self): # her testinin bitiminde çalışacak fonk
        self.driver.quit()

    def getData():
        excel = openpyxl.load_workbook("data/valid_login.xlsx")
        sheet = excel["Sheet1"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))

        return data   
    
    def test_passing_empty(self):
        usernameInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        usernameInput.send_keys("")
        passwordInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        passwordInput.send_keys("")
        loginButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.LOGIN_BUTTON)))
        loginButton.click()
        errorMessage = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.ERROR_MESSAGE_XPATH)))
        assert errorMessage.text == "Epic sadface: Username is required"
        

#Sadece şifre alanı boş geçildiğinde uyarı mesajı olarak "Epic sadface: Password is required" gösterilmelidir.
        
    def test_invalid_password(self):
        usernameInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        usernameInput.send_keys("locked_out_user")
        passwordInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        passwordInput.send_keys("")
        loginButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.LOGIN_BUTTON)))
        loginButton.click()
        errorMessage = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.ERROR_MESSAGE_XPATH)))
        assert errorMessage.text == "Epic sadface: Password is required"
        

#Kullanıcı adı "locked_out_user" şifre alanı "secret_sauce" gönderildiğinde "Epic sadface: Sorry, this user has been locked out." mesajı gösterilmelidir.
        
    @pytest.mark.parametrize("Username, Password", [("locked_out_user", "secret_sauce")]) 
    def test_invalid_login(self, Username, Password):
        usernameInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        usernameInput.send_keys(Username)
        passwordInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        passwordInput.send_keys(Password)
        loginButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.LOGIN_BUTTON)))
        loginButton.click()
        errorMessage = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.ERROR_MESSAGE_XPATH)))
        assert errorMessage.text == "Epic sadface: Sorry, this user has been locked out."
        

#Kullanıcı adı "standard_user" şifre "secret_sauce" gönderildiğinde kullanıcı "/inventory.html" sayfasına gönderilmelidir. 
#Giriş yapıldıktan sonra kullanıcıya gösterilen ürün sayısı "6" adet olmalıdır. 
    
    @pytest.mark.parametrize("Username, Password", getData())
    def test_product_control(self, Username, Password):
        usernameInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        usernameInput.send_keys(Username)
        passwordInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        passwordInput.send_keys(Password)
        loginButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.LOGIN_BUTTON)))
        loginButton.click()
        current_url = self.driver.current_url 
        expected_url = c.INVENTORY_URL
        if (current_url == expected_url):
            first_test = True
        else:
            first_test = False
        product = WebDriverWait(self.driver,2).until(ec.visibility_of_all_elements_located((By.CLASS_NAME, c.PRODUCTS_CONTROL)))
        second_test = len(product) == 6
        if (first_test == True and second_test == True):
            assert True
        else:
            assert False

#Sitedeki seçtiğin ilk ürüne tıkladığında detay sayfasına gidiliyor mu?  
                   
class Test_Sauce_Demo_Product_Test:
        #prefix => test_ 
    def setup_method(self): #her test başlangıcında çalışacak fonk
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window() 
    
    def teardown_method(self): # her testinin bitiminde çalışacak fonk
        self.driver.quit()
    
    def test_product_details(self):
        usernameInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        passwordInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.LOGIN_BUTTON)))
        loginButton.click()
        current_url = self.driver.current_url 
        expected_url = c.INVENTORY_URL
        if (current_url == expected_url):
            first_test = True
        else:
            first_test = False
        first_product = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.FIRST_PRODUCT)))
        first_product.click()
        current_url = self.driver.current_url 
        expected_url = c.FIRST_PRODUCT_DETAIL_URL
        if (current_url == expected_url):
            assert True
        else:
            assert False

#Sayfada üç adet ürün seçildiğinde sepet ikonu üzerindeki üç sayısıyla seçilen ürün adeti sayısı eşit mi?

    def test_product_quantity(self):
        usernameInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        passwordInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.LOGIN_BUTTON)))
        loginButton.click()
        current_url = self.driver.current_url 
        expected_url = c.INVENTORY_URL
        if (current_url == expected_url):
            assert True
        else:
            assert False
        addToCart1 = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.ADD_TO_CART_1)))
        addToCart1.click()
        addToCart2 = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.ADD_TO_CART_2)))
        addToCart2.click()
        addToCart3 = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.ADD_TO_CART_3)))
        addToCart3.click()
        cart_badge = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.CLASS_NAME, c.CART_BADGE_TEXT))).text
        if (cart_badge =="3"):
            assert True
        else:
            assert False
        

#Siteden hatasız bir şekilde çıkış yapılıyor mu?     

class Test_Sauce_Demo_Logout_Test:              
        #prefix => test_ 
    def setup_method(self): #her test başlangıcında çalışacak fonk
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window() 
    
    def teardown_method(self): # her testinin bitiminde çalışacak fonk
        self.driver.quit()
    
    def test_logout(self):
        usernameInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        passwordInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.LOGIN_BUTTON)))
        loginButton.click()
        current_url = self.driver.current_url 
        expected_url = c.INVENTORY_URL
        if (current_url == expected_url):
            first_test = True
        else:
            first_test = False
        menuButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.MENU_BUTTON)))
        menuButton.click()
        logoutButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.LOGOUT_BUTTON)))
        logoutButton.click()
        current_url_2 = self.driver.current_url
        expected_url = c.BASE_URL
        assert current_url_2 == expected_url
        
#Alfabetik sıraya göre filtreleme yapılıyor mu?
        
class Test_Sauce_Demo_Filter_Test:
        #prefix => test_ 
    def setup_method(self): #her test başlangıcında çalışacak fonk
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window() 
    
    def teardown_method(self): # her testinin bitiminde çalışacak fonk
        self.driver.quit()
    
    def test_filter(self):
        usernameInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        passwordInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.LOGIN_BUTTON)))
        loginButton.click()
        current_url = self.driver.current_url 
        expected_url = c.INVENTORY_URL
        if (current_url == expected_url):
            assert True
        else:
            assert False
        filterButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.CLASS_NAME, c.FILTER_BUTTON)))
        filterButton.click()
        zToa = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.Z_TO_A_FILTER)))
        zToa.click()
        first_product = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.FIRST_PRODUCT))).text
        second_product = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.SECOND_PRODUCT))).text
        third_product = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.THIRD_PRODUCT))).text
        forth_product = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.FORTH_PRODUCT))).text
        fifth_product = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.FIFTH_PRODUCT))).text
        sixth_product = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.ID, c.SIXTH_PRODUCT))).text
        list = [first_product, second_product, third_product, forth_product, fifth_product, sixth_product]
        reverse_list = sorted(list, reverse=True)
        assert reverse_list
    

